"""
讀取 Google Sheet 採購表，提取商品清單。
支援從 Google Sheets 匯出的 .xlsx 檔案，自動提取超連結 URL。
"""
import re
from dataclasses import dataclass
from pathlib import Path

import httpx
import openpyxl
from loguru import logger


@dataclass
class SheetProduct:
    """採購表中的一筆商品資料。"""
    row_number: int
    date: str               # A: 填表日期
    product_name: str       # B: 商品 or 品牌名稱
    reference_url: str      # C: 商品參考網址（hyperlink）
    monthly_sales: str      # D: 月銷
    category: str           # E: 分類
    style_desc: str         # F: 約略預計進貨款式與數量
    variant_count: int      # G: 款式顏色乘積
    qty_per_unit: int       # H: 預計每單位數量
    purchase_url: str       # I: 進貨網址（1688 URL，hyperlink）
    supplier_name: str      # J: 廠商名稱
    selling_price: float    # L: 商品售價（TWD）
    notes: str              # M: 備註
    exchange_rate: float    # Row 1 B 的匯率

    @property
    def item_id(self) -> str | None:
        """從 1688 URL 提取 item_id。"""
        url = self.purchase_url or self.reference_url
        if not url:
            return None
        m = re.search(r'offer/(\d+)\.html', url)
        return m.group(1) if m else None

    @property
    def url_1688(self) -> str:
        """取得 1688 商品 URL（優先 purchase_url）。"""
        return self.purchase_url or self.reference_url or ""


def download_sheet(
    sheet_id: str,
    gid: str,
    output_path: Path,
) -> Path:
    """
    從 Google Sheets 匯出 URL 下載 .xlsx。

    注意：需要 Sheet 有「知道連結的人皆可檢視」權限。
    """
    url = (
        f"https://docs.google.com/spreadsheets/d/{sheet_id}"
        f"/export?format=xlsx&gid={gid}"
    )
    logger.info(f"下載採購表: {url}")

    with httpx.Client(follow_redirects=True, timeout=30) as client:
        resp = client.get(url)
        resp.raise_for_status()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(resp.content)
    logger.info(f"採購表已下載: {output_path} ({len(resp.content)} bytes)")
    return output_path


def _extract_url(cell) -> str:
    """從 openpyxl cell 提取超連結 URL。"""
    # 優先用 hyperlink 屬性
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target

    # 有些 Google Sheet 匯出的超連結存在 value 本身
    val = str(cell.value or "")
    if val.startswith("http"):
        return val

    return ""


def _safe_int(val, default: int = 0) -> int:
    """安全轉換為 int。"""
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def _safe_float(val, default: float = 0.0) -> float:
    """安全轉換為 float。"""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def read_procurement_sheet(xlsx_path: Path) -> tuple[float, list[SheetProduct]]:
    """
    讀取採購表 .xlsx，回傳 (匯率, 商品清單)。

    欄位對應：
    Row 1: A=匯率, B=匯率值
    Row 2: Headers
    Row 3+: 資料

    A=填表日期, B=商品名稱, C=參考網址, D=月銷, E=分類,
    F=款式描述, G=款式數, H=每單位數量, I=進貨網址,
    J=廠商名稱, K=商品售價, L=備註
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # 讀取匯率（Row 1, Cell B1）
    exchange_rate = _safe_float(ws.cell(1, 2).value, 4.81)
    logger.info(f"匯率: {exchange_rate}")

    products = []
    for row_num in range(3, ws.max_row + 1):
        # 跳過空行
        product_name = ws.cell(row_num, 2).value
        if not product_name:
            continue

        product = SheetProduct(
            row_number=row_num,
            date=str(ws.cell(row_num, 1).value or ""),
            product_name=str(product_name).strip(),
            reference_url=_extract_url(ws.cell(row_num, 3)),
            monthly_sales=str(ws.cell(row_num, 4).value or ""),
            category=str(ws.cell(row_num, 5).value or "").strip(),
            style_desc=str(ws.cell(row_num, 6).value or "").strip(),
            variant_count=_safe_int(ws.cell(row_num, 7).value, 1),
            qty_per_unit=_safe_int(ws.cell(row_num, 8).value, 5),
            purchase_url=_extract_url(ws.cell(row_num, 9)),
            supplier_name=str(ws.cell(row_num, 10).value or "").strip(),
            selling_price=_safe_float(ws.cell(row_num, 11).value, 0),
            notes=str(ws.cell(row_num, 12).value or "").strip(),
            exchange_rate=exchange_rate,
        )

        # 至少要有 URL 才有意義
        if product.item_id:
            products.append(product)
            logger.debug(f"Row {row_num}: {product.product_name} -> {product.item_id}")
        else:
            logger.warning(f"Row {row_num}: {product.product_name} - 無有效 1688 URL，跳過")

    wb.close()
    logger.info(f"共讀取 {len(products)} 筆有效商品")
    return exchange_rate, products
